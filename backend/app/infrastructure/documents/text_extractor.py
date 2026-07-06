"""Plain-text extraction from uploaded documents.

Framework-agnostic: no FastAPI, SQLAlchemy, or AI/RAG imports. Given a file
path, returns the extracted plain text. Each format's parsing library is
imported lazily inside its own extraction function, so a deployment only
needs the dependency for the formats it actually processes.

Migrated from the Hemaya reference project's ``backend/text_extractor.py``.
Unlike the Hemaya original, extraction failures raise a typed exception
instead of returning a sentinel string -- this lets callers distinguish a
genuine empty document from a parsing failure without string-sniffing.
"""

import asyncio
import logging
from pathlib import Path
from typing import Callable, Dict

logger = logging.getLogger(__name__)


class TextExtractionError(Exception):
    """Raised when a document's text cannot be extracted."""


class UnsupportedFileTypeError(TextExtractionError):
    """Raised when the file extension has no registered extraction strategy."""


async def extract_text(file_path: Path) -> str:
    """Extract plain text from ``file_path`` based on its file extension.

    The (blocking) parsing work runs in a worker thread so callers on an
    asyncio event loop -- e.g. a future FastAPI upload endpoint -- are not
    blocked while a large PDF or workbook is parsed.

    Raises:
        UnsupportedFileTypeError: the extension has no extraction strategy.
        TextExtractionError: the file exists but could not be parsed.
    """
    return await asyncio.to_thread(_extract_text_sync, file_path)


def _extract_text_sync(file_path: Path) -> str:
    """Synchronous extraction dispatch, used directly by tests."""
    extension = file_path.suffix.lower().lstrip(".")
    strategy = _STRATEGIES.get(extension)
    if strategy is None:
        raise UnsupportedFileTypeError(
            f"No text extraction strategy registered for '.{extension}' files"
        )
    try:
        return strategy(file_path)
    except TextExtractionError:
        raise
    except Exception as exc:
        logger.exception("Text extraction failed for %s", file_path)
        raise TextExtractionError(
            f"Failed to extract text from '{file_path.name}': {exc}"
        ) from exc


def _extract_pdf(file_path: Path) -> str:
    """Extract text from a PDF, page by page, using PyMuPDF."""
    import pymupdf

    doc = pymupdf.open(str(file_path))
    try:
        return "\n".join(page.get_text() for page in doc.pages())
    finally:
        doc.close()


def _extract_docx(file_path: Path) -> str:
    """Extract text from a Word document, one paragraph per line."""
    from docx import Document

    document = Document(str(file_path))
    return "\n".join(p.text for p in document.paragraphs if p.text.strip())


def _extract_xlsx(file_path: Path) -> str:
    """Extract text from an Excel workbook, tab-separated per row."""
    from openpyxl import load_workbook

    workbook = load_workbook(str(file_path), read_only=True, data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                lines.append(row_text)
    return "\n".join(lines)


def _extract_txt(file_path: Path) -> str:
    """Extract text from a plain-text file, auto-detecting its encoding."""
    import chardet

    raw = file_path.read_bytes()
    detected = chardet.detect(raw)
    encoding = detected.get("encoding") or "utf-8"
    return raw.decode(encoding, errors="replace")


_STRATEGIES: Dict[str, Callable[[Path], str]] = {
    "pdf": _extract_pdf,
    "docx": _extract_docx,
    "xlsx": _extract_xlsx,
    "xls": _extract_xlsx,
    "txt": _extract_txt,
}
